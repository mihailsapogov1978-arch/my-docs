import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from openpyxl import load_workbook


def categorize_issue(topic, description):
    """
    Категоризирует заявку по проблемным направлениям на основе темы и описания
    """
    text = str(topic).lower() + " " + str(description).lower()

    # 1. Технические проблемы
    tech_keywords = [
        'запуск', 'обновл', 'права доступа', 'доступ', 'парол', 'вход', 'авторизац',
        'завис', 'тормоз', 'ошибка', 'не открывает', 'не заходит', 'эп', 'эцп',
        'подпис', 'сертификат', 'технич', 'сбой', 'работает', 'не работает',
        'версия', 'установк', 'инсталл', 'клиент', 'веб', 'браузер', 'интернет'
    ]

    # 2. Бухгалтерский учет и налогообложение
    accounting_keywords = [
        'бухгалтер', 'учет', 'операц', 'проводк', 'счет', 'дебет', 'кредит',
        'зарплат', 'заработн', 'оплат', 'выплат', 'начисл', 'удержан', 'преми',
        'налог', 'ндфл', 'страхов', 'взнос', 'фсс', 'пфр', 'период', 'закрыт',
        'отчетн', 'баланс', 'прибыл', 'убыток', 'амортиз', 'основн', 'касс',
        'банк', 'платеж', 'поручен', 'аванс'
    ]

    # 3. Взаимодействие с внешними системами
    external_keywords = [
        'ркс', 'еис', 'гис имущество', 'банк', 'платеж', 'поручен', 'выписк',
        'казначей', 'фин', 'бюджет', 'госзакуп', 'закупк', 'тендер', 'конкурс',
        'электрон', 'эд', 'документооборот', 'архив', 'хранен', 'офд', 'оператор'
    ]

    # 4. Формирование регламентированной отчетности
    reporting_keywords = [
        'отчетност', 'фнс', 'сфр', 'росстат', 'астрал', 'контур', 'такском',
        'сзв', 'рсв', '4-фсс', '2-ндфл', '6-ндфл', 'бухгалтерск', 'финансов',
        'статистич', 'декларац', 'налогов'
    ]

    # 5. Кадровый учет
    hr_keywords = [
        'кадр', 'сотрудник', 'работник', 'персонал', 'гис тк', 'сзв-тд',
        'труд', 'больничн', 'отпуск', 'отгул', 'командиров', 'стаж', 'тк',
        'трудовой', 'договор', 'контракт', 'прием', 'увольнен', 'перевод',
        'гис ку', 'гис еску', 'единый', 'реестр'
    ]

    # 6. Личный кабинет подотчетного лица
    personal_keywords = [
        'личный кабинет', 'лк', 'подотчетн', 'авансов', 'отчет', 'командировоч',
        'подотчетное лицо', 'физ лицо', 'сотруднический'
    ]

    # 7. Дополнительный функционал и сервисы
    additional_keywords = [
        'аналитик', 'планирован', 'бюджетирован', 'консолидац', 'мониторинг',
        'контрол', 'апи', 'интеграц', 'api', 'веб-сервис', 'сервис', 'модуль',
        'дополнительн', 'расширен', 'новый функционал'
    ]

    # 8. Справочники и классификаторы
    directory_keywords = [
        'справочник', 'классификатор', 'инн', 'кпп', 'огрн', 'задвоен',
        'дубликат', 'повтор', 'аналитик', 'иерархи', 'структур', 'подразделен',
        'организац', 'контрагент', 'поставщик', 'покупатель', 'сотрудничеств',
        'номенклатур', 'материал', 'ос', 'основные средства', 'мног'
    ]

    categories = {
        'Технические проблемы': any(keyword in text for keyword in tech_keywords),
        'Бухгалтерский учет': any(keyword in text for keyword in accounting_keywords),
        'Внешние системы': any(keyword in text for keyword in external_keywords),
        'Регламентированная отчетность': any(keyword in text for keyword in reporting_keywords),
        'Кадровый учет': any(keyword in text for keyword in hr_keywords),
        'Личный кабинет': any(keyword in text for keyword in personal_keywords),
        'Дополнительный функционал': any(keyword in text for keyword in additional_keywords),
        'Справочники': any(keyword in text for keyword in directory_keywords)
    }

    for category, matches in categories.items():
        if matches:
            return category

    return 'Другое'


def parse_excel_to_md(excel_path, md_path):
    """
    Парсит Excel файл с заявками и ПЕРЕЗАПИСЫВАЕТ отчет в markdown файл
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        print(f"📊 Загружаем файл: {excel_path}")
        print(f"   Всего строк в файле: {ws.max_row}")

        header_row = None
        for i in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row=i, column=1).value
            if cell_value in ['Номер', 1, '1']:
                header_row = i
                print(f"✅ Найдены заголовки на строке {i}")
                break

        if header_row:
            df = pd.read_excel(excel_path, header=header_row - 1)
        else:
            df = pd.read_excel(excel_path, header=3)
            print("⚠️ Заголовки не найдены, используем строку 4")

        print(f"✅ Данные загружены. Строк: {len(df)}, Колонок: {len(df.columns)}")

    except Exception as e:
        print(f"❌ Ошибка при чтении файла: {e}")
        return None

    df.columns = [str(col).strip() for col in df.columns]

    column_mapping = {}
    column_patterns = {
        'Номер': ['номер', '№', '1'],
        'Дата': ['дата', 'время', 'регистрац', '2'],
        'Инициатор': ['инициатор', 'заявки', '3'],
        'Статус': ['статус', 'обращен', '4'],
        'Тема': ['тема', '5'],
        'Описание': ['описание', '6'],
        'Исполнитель': ['исполнитель', '7'],
        'Решение': ['решение', '8'],
        'Услуга': ['услуга', '9']
    }

    print("\n🔍 Определяем колонки:")
    for col_name, patterns in column_patterns.items():
        for pattern in patterns:
            for df_col in df.columns:
                if pattern in str(df_col).lower():
                    column_mapping[col_name] = df_col
                    print(f"   • {col_name}: '{df_col}'")
                    break
            if col_name in column_mapping:
                break

    report_time = datetime.now().strftime("%d.%m.%Y %H:%M")
    total_requests = len(df)
    print(f"\n📊 Всего заявок: {total_requests}")

    status_counts = pd.Series(dtype='object')
    if 'Статус' in column_mapping:
        status_col = column_mapping['Статус']
        status_series = df[status_col].dropna()
        if len(status_series) > 0:
            status_series = status_series.astype(str).str.strip()
            status_counts = status_series.value_counts()

    org_stats = {}
    gov_keywords = ['ГОСУДАРСТВЕНН', 'ГКУ', 'ГУ', 'ГБУ', 'ГАУ', 'ГОСУДАРСТВЕННОЕ', 'ГОСУДАРСТВЕННЫЙ']
    cb_keyword = 'ЦЕНТРАЛИЗОВАННАЯ БУХГАЛТЕРИЯ ОРГАНОВ ГОСУДАРСТВЕННОЙ ВЛАСТИ'

    iogv_keywords = [
        'ДЕПАРТАМЕНТ ОБРАЗОВАНИЯ ЯМАЛО-НЕНЕЦКОГО',
        'ДЕПАРТАМЕНТ КУЛЬТУРЫ ЯМАЛО-НЕНЕЦКОГО',
        'ДЕПАРТАМЕНТ ПО ДЕЛАМ КОРЕННЫХ',
        'ДЕПАРТАМЕНТ ИНФОРМАЦИОННЫХ ТЕХНОЛОГИЙ',
        'ДЕПАРТАМЕНТ МОЛОДЁЖНОЙ ПОЛИТИКИ',
        'ДЕПАРТАМЕНТ ЭКОНОМИКИ ЯМАЛО-НЕНЕЦКОГО',
        'ДЕПАРТАМЕНТ ПО ФИЗИЧЕСКОЙ КУЛЬТУРЕ',
        'ДЕПАРТАМЕНТ ГРАЖДАНСКОЙ ЗАЩИТЫ',
        'СЛУЖБА ЗАПИСИ АКТОВ ГРАЖДАНСКОГО СОСТОЯНИЯ',
        'ДЕПАРТАМЕНТ ВНУТРЕННЕЙ ПОЛИТИКИ',
        'ДЕПАРТАМЕНТ ПРИРОДНЫХ РЕСУРСОВ',
        'ДЕПАРТАМЕНТ ВНЕШНИХ СВЯЗЕЙ',
        'ДЕПАРТАМЕНТ ТРАНСПОРТА И ДОРОЖНОГО ХОЗЯЙСТВА',
        'ДЕПАРТАМЕНТ ЗДРАВООХРАНЕНИЯ ЯМАЛО-НЕНЕЦКОГО',
        'СЛУЖБА ВЕТЕРИНАРИИ ЯМАЛО-НЕНЕЦКОГО',
        'ДЕПАРТАМЕНТ СТРОИТЕЛЬСТВА И ЖИЛИЩНОЙ ПОЛИТИКИ',
        'ДЕПАРТАМЕНТ АГРОПРОМЫШЛЕННОГО КОМПЛЕКСА',
        'ДЕПАРТАМЕНТ РЕГИОНАЛЬНОЙ БЕЗОПАСНОСТИ',
        'УПРАВЛЕНИЕ ДЕЛАМИ ПРАВИТЕЛЬСТВА ЯМАЛО-НЕНЕЦКОГО'
    ]

    current_org = None
    for idx, row in df.iterrows():
        org_value = row.iloc[0] if len(df.columns) > 0 else None
        if pd.notna(org_value) and isinstance(org_value, str):
            org_str = str(org_value).strip()
            if org_str and len(org_str) > 5:
                current_org = org_str

        if current_org:
            if current_org not in org_stats:
                org_stats[current_org] = {'total': 0, 'resolved': 0, 'in_work': 0}

            status = ''
            if 'Статус' in column_mapping and pd.notna(row.get(column_mapping['Статус'])):
                status = str(row[column_mapping['Статус']]).strip().lower()

            org_stats[current_org]['total'] += 1
            if 'закрыт' in status or 'решен' in status:
                org_stats[current_org]['resolved'] += 1
            elif any(word in status for word in ['назначен', 'в работе', 'зарегистрирован']):
                org_stats[current_org]['in_work'] += 1

    print(f"✅ Найдено организаций: {len(org_stats)}")

    gov_orgs = []
    mun_orgs = []
    iogv_orgs = []

    gov_total = gov_resolved = gov_in_work = 0
    mun_total = mun_resolved = mun_in_work = 0
    iogv_total = iogv_resolved = iogv_in_work = 0

    for org_name, stats in org_stats.items():
        org_upper = org_name.upper()
        is_cb = cb_keyword in org_upper
        is_iogv_dept = any(keyword in org_upper for keyword in iogv_keywords)

        if is_cb or is_iogv_dept:
            iogv_orgs.append((org_name, stats['total']))
            iogv_total += stats['total']
            iogv_resolved += stats['resolved']
            iogv_in_work += stats['in_work']
        elif any(keyword in org_upper for keyword in gov_keywords):
            gov_orgs.append((org_name, stats['total']))
            gov_total += stats['total']
            gov_resolved += stats['resolved']
            gov_in_work += stats['in_work']
        else:
            mun_orgs.append((org_name, stats['total']))
            mun_total += stats['total']
            mun_resolved += stats['resolved']
            mun_in_work += stats['in_work']

    gov_orgs.sort(key=lambda x: x[1], reverse=True)
    mun_orgs.sort(key=lambda x: x[1], reverse=True)
    iogv_orgs.sort(key=lambda x: x[1], reverse=True)

    print(f"📊 Типы учреждений (новая классификация):")
    print(f"   • Государственные учреждения: {gov_total} заявок ({len(gov_orgs)} организаций)")
    print(f"   • Муниципальные учреждения: {mun_total} заявок ({len(mun_orgs)} организаций)")
    print(f"   • ИОГВ ЯНАО: {iogv_total} заявок ({len(iogv_orgs)} организаций)")

    issue_categories = {}
    if 'Тема' in column_mapping and 'Описание' in column_mapping:
        theme_col = column_mapping['Тема']
        desc_col = column_mapping['Описание']
        print("\n🔍 Анализируем проблемные направления...")
        sample_size = min(200, len(df))
        for idx in range(sample_size):
            row = df.iloc[idx]
            topic = row[theme_col] if pd.notna(row.get(theme_col)) else ''
            description = row[desc_col] if pd.notna(row.get(desc_col)) else ''
            category = categorize_issue(topic, description)
            issue_categories[category] = issue_categories.get(category, 0) + 1
        print(f"✅ Проанализировано {sample_size} заявок")

    sorted_issues = sorted(issue_categories.items(), key=lambda x: x[1], reverse=True)

    md_content = f"""# 📊 Отчет по заявкам

**Обновлено:** {report_time}  
**Всего заявок:** {total_requests}

---

## 📈 Заявки ГМУ и ИОГВ за период с 5.01.2026 по 5.02.2026

### Общая статистика

| Тип учреждения | Всего заявок | Решено | В работе |
|----------------|--------------|--------|----------|
| Государственные учреждения | {gov_total} | {gov_resolved} | {gov_in_work} |
| Муниципальные учреждения | {mun_total} | {mun_resolved} | {mun_in_work} |
| ИОГВ ЯНАО | {iogv_total} | {iogv_resolved} | {iogv_in_work} |

---

### Анализ проблемных направлений

| № | Категория проблемы | Количество заявок | Доля |
|---|-------------------|-------------------|------|
"""

    if sorted_issues:
        sample_total = sum(issue_categories.values())
        for i, (category, count) in enumerate(sorted_issues, 1):
            percentage = round((count / sample_total) * 100, 1) if sample_total > 0 else 0
            md_content += f"| {i} | {category} | {count} | {percentage}% |\n"
    else:
        md_content += "| 1 | *Нет данных для анализа* | 0 | 0% |\n"

    md_content += "\n---\n\n"

    # ✅ Безопасное создание директории (только если путь содержит подпапку)
    output_dir = os.path.dirname(md_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(md_content)

    print(f"\n✅ Основной отчет успешно ПЕРЕЗАПИСАН: {md_path}")
    print(f"📊 Проанализировано заявок: {total_requests}")
    print(f"🏢 Государственные учреждения: {gov_total} заявок ({len(gov_orgs)} организаций)")
    print(f"🏢 Муниципальные учреждения: {mun_total} заявок ({len(mun_orgs)} организаций)")
    print(f"🏛️ ИОГВ ЯНАО: {iogv_total} заявок ({len(iogv_orgs)} организаций)")

    if iogv_orgs:
        print(f"\n🔍 Организации ИОГВ ЯНАО ({len(iogv_orgs)} наименований):")
        for i, (org_name, count) in enumerate(iogv_orgs, 1):
            print(f"   {i:3}. {org_name[:80]}: {count} заявок")

    return {
        'total_requests': total_requests,
        'gov_total': gov_total,
        'gov_resolved': gov_resolved,
        'gov_in_work': gov_in_work,
        'gov_orgs': gov_orgs,
        'mun_total': mun_total,
        'mun_resolved': mun_resolved,
        'mun_in_work': mun_in_work,
        'mun_orgs': mun_orgs,
        'iogv_total': iogv_total,
        'iogv_resolved': iogv_resolved,
        'iogv_in_work': iogv_in_work,
        'iogv_orgs': iogv_orgs,
        'status_counts': dict(status_counts),
        'issue_categories': dict(sorted_issues)
    }


def generate_zayavky_report():
    """
    Читает данные из файла zayavky.xlsx (лист "Заявки")
    и добавляет их в конец файла zayavky.md
    """
    print("\n📖 Читаем дополнительный файл: zayavky.xlsx")

    try:
        file_path = "zayavky.xlsx"
        if not os.path.exists(file_path):
            print(f"❌ Файл {file_path} не найден!")
            return False

        try:
            df = pd.read_excel(file_path, sheet_name="Заявки", dtype=str)
            print("✅ Лист 'Заявки' прочитан успешно")
        except:
            print("⚠️ Лист 'Заявки' не найден, читаем первый лист")
            df = pd.read_excel(file_path, dtype=str)

        required_columns = ['Статус', 'Инициатор']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"⚠️ В файле отсутствуют колонки: {missing_columns}")
            for col in required_columns:
                for df_col in df.columns:
                    if col.lower() in df_col.lower():
                        df.rename(columns={df_col: col}, inplace=True)
                        print(f"   • Переименовали '{df_col}' в '{col}'")
                        break

        total = len(df)
        print(f"📊 Всего заявок в zayavky.xlsx: {total}")

        resolved = in_work = closed = 0
        if 'Статус' in df.columns:
            df['Статус'] = df['Статус'].astype(str).str.strip()
            resolved = df[df['Статус'].str.contains('Решен', case=False, na=False)].shape[0]
            in_work = df[df['Статус'].str.contains('В работе', case=False, na=False)].shape[0]
            closed = df[df['Статус'].str.contains('Закрыт', case=False, na=False)].shape[0]
            print(f"   • Решено: {resolved}")
            print(f"   • В работе: {in_work}")
            print(f"   • Закрыто: {closed}")
        else:
            print("⚠️ Колонка 'Статус' не найдена для статистики")

        top_3_table = ""
        if 'Инициатор' in df.columns:
            df['Инициатор'] = df['Инициатор'].astype(str).str.strip()
            initiator_counts = df['Инициатор'].value_counts().head(3)
            if not initiator_counts.empty:
                top_3_table = "| Инициатор | Количество заявок |\n|:---|---:|\n"
                for name, count in initiator_counts.items():
                    name_display = str(name)[:30] + ('...' if len(str(name)) > 30 else '')
                    top_3_table += f"| {name_display} | {count} |\n"
                print(f"✅ Найдено {len(initiator_counts)} активных инициаторов")
            else:
                top_3_table = "*Нет данных об инициаторах*\n"
                print("⚠️ Нет данных об инициаторах")
        else:
            top_3_table = "*Колонка 'Инициатор' не найдена*\n"
            print("⚠️ Колонка 'Инициатор' не найдена")

        md_content = f"""
## 📊 Заявки ЦБ ОГВ ЯНАО за прошлую неделю

### 1. Общая статистика

**Всего заявок:** {total}  
**Из них:**
- Решено: {resolved}
- В работе: {in_work} 
- Закрыто: {closed}
- Ожидают обработки: {total - resolved - in_work - closed}

### 2. Топ-3 активных инициаторов

{top_3_table}
"""
        return md_content

    except Exception as e:
        print(f"❌ Ошибка при чтении файла zayavky.xlsx: {e}")
        return False


if __name__ == "__main__":
    excel_file = "zayavky_all.xlsx"
    md_file = "zayavky.md"  # ✅ Теперь точно без docs/

    print("=" * 60)
    print("🚀 ЗАПУСК ГЕНЕРАЦИИ ОТЧЕТА")
    print("=" * 60)

    if os.path.exists(md_file):
        print(f"🗑️ Удаляем старый файл: {md_file}")
        os.remove(md_file)

    if os.path.exists(excel_file):
        stats = parse_excel_to_md(excel_file, md_file)
        if stats:
            additional_content = generate_zayavky_report()
            if additional_content:
                with open(md_file, 'a', encoding='utf-8') as f:
                    f.write(additional_content)
                print("\n✅ Дополнительные данные из zayavky.xlsx добавлены в конец файла")
            else:
                print("\n⚠️ Не удалось добавить данные из zayavky.xlsx")

            print("\n" + "=" * 60)
            print("✅ ФИНАЛЬНЫЙ ОТЧЕТ УСПЕШНО СОЗДАН")
            print("=" * 60)
            print(f"📋 Основные показатели:")
            print(f"   • Всего заявок: {stats['total_requests']}")
            print(f"   • Государственные учреждения: {stats['gov_total']} заявок ({len(stats['gov_orgs'])} организаций)")
            print(f"     (Решено: {stats['gov_resolved']}, В работе: {stats['gov_in_work']})")
            print(f"   • Муниципальные учреждения: {stats['mun_total']} заявок ({len(stats['mun_orgs'])} организаций)")
            print(f"     (Решено: {stats['mun_resolved']}, В работе: {stats['mun_in_work']})")
            print(f"   • ИОГВ ЯНАО: {stats['iogv_total']} заявок ({len(stats['iogv_orgs'])} организаций)")
            print(f"     (Решено: {stats['iogv_resolved']}, В работе: {stats['iogv_in_work']})")
            print(f"\n📄 Файл создан: {md_file}")
    else:
        print(f"❌ Основной файл {excel_file} не найден!")
        if os.path.exists("zayavky.xlsx"):
            print("⚠️ Но найден файл zayavky.xlsx — создаём отчёт только из него")
            report_time = datetime.now().strftime("%d.%m.%Y %H:%M")
            md_content = f"""# 📊 Отчет по заявкам

**Обновлено:** {report_time}

"""
            additional_content = generate_zayavky_report()
            if additional_content:
                md_content += additional_content
                with open(md_file, 'w', encoding='utf-8') as f:
                    f.write(md_content)
                print(f"\n✅ Отчет создан только из zayavky.xlsx: {md_file}")
            else:
                print("❌ Не удалось создать отчет")